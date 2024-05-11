#!/usr/bin/env python
# coding: utf-8

# In[2]:


from openpyxl import load_workbook  # Import load_workbook directly
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import sys

application_path = os.path.dirname(sys.executable)

#Putting together add-chart, apply-formula and format scripts (inputs: pivot-table + month; outputs: charts, formula and format)
month = input('Introduce month: ')

# Load the workbook and get the 'Report' sheet
input_path = os.path.join(application_path, 'pivot_table.xlsx')
wb = load_workbook(input_path)
sheet = wb['Report']

# Get the dimensions of the active sheet
min_column = sheet.min_column
max_column = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

# -- 1. Add Chart
# Create a new BarChart object
chart = BarChart()

# Define the data range for the chart
data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)

# Define the categories (x-axis labels) for the chart
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)

# Add data and categories to the chart
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Add the chart to the 'Report' sheet at cell B12
sheet.add_chart(chart, "B12")

# Set the title of the chart
chart.title = "Sales by Product Line"

# Apply a pre-defined style to the chart
chart.style = 2

# -- 2. Apply Formula
# Iterate through each column from min_column+1 to max_column+1
for i in range(min_column+1, max_column+1):
    # Get the letter corresponding to the current column index
    letter = get_column_letter(i)

    # Calculate the sum formula for the current column and insert it into the cell below the column data
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1} : {letter}{max_row})'

    # Apply 'Currency' style to the cell
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

sheet[f'{get_column_letter(min_column)}{max_row+1}'] = 'Total'

# -- 3. Add Format
# Set the values of cell A1 and A2
sheet['A1'] = 'Sales Report'
sheet['A2'] = month

# Apply formatting to cell A1 and A2
sheet['A1'].font = Font('Aptos', bold=True, size=20)
sheet['A2'].font = Font('Aptos', bold=True, size=13)

# Save the workbook with the chart to a new Excel file
output_path = os.path.join(application_path, f'report_{month}.xlsx')
wb.save(output_path)



# In[ ]:




